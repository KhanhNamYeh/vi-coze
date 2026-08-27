"""XLSX -> Markdown; tên feature lấy trực tiếp từ header của workbook."""

from pathlib import Path

from ..settings import PreprocessSettings


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _heading(level: int, value: object) -> str:
    return f"{'#' * level} {_text(value).replace(chr(10), ' ')}"


def _id_index(headers: list[str], column: int | str) -> int:
    if isinstance(column, int):
        if not 0 <= column < len(headers):
            raise ValueError(f"excel_id_column={column} nằm ngoài {len(headers)} cột")
        return column
    try:
        return headers.index(column)
    except ValueError as error:
        raise ValueError(f"không có cột ID '{column}'; có: {', '.join(headers)}") from error


def to_markdown(source: str | Path, settings: PreprocessSettings) -> str:
    import openpyxl

    path = Path(source)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    selected = settings.excel_sheets or workbook.sheetnames
    missing = [name for name in selected if name not in workbook.sheetnames]
    if missing:
        workbook.close()
        raise ValueError(f"{path.name}: không có sheet {', '.join(missing)}")

    output: list[str] = []
    excluded = {value.strip().casefold() for value in settings.excel_exclude_features}
    for sheet_name in selected:
        rows = workbook[sheet_name].iter_rows(values_only=True)
        first = next(rows, None)
        if first is None:
            continue
        headers = [_text(value) for value in first]
        if not any(headers):
            continue
        id_index = _id_index(headers, settings.excel_id_column)
        output.append(_heading(1, sheet_name))

        for row in rows:
            values = list(row) + [None] * max(0, len(headers) - len(row))
            record_id = _text(values[id_index])
            if not record_id:
                continue
            output.extend(["", _heading(settings.record_heading_level, record_id)])
            for index, feature in enumerate(headers):
                if index == id_index or not feature or feature.casefold() in excluded:
                    continue
                output.extend([
                    "",
                    _heading(settings.feature_heading_level, feature),
                    _text(values[index]),
                ])
    workbook.close()

    markdown = "\n".join(output).strip()
    if not markdown:
        raise ValueError(f"{path.name}: không có record Excel hợp lệ")
    return markdown + "\n"
