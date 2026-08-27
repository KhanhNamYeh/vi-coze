"""Đánh giá graph trên split dev/test; không nằm trong offline pipeline."""

from __future__ import annotations

import copy
import json
import re

from ..offline.graph import load as load_graph
from ..settings import Settings, load_settings


def combine(
    doc_ids: list[str],
    *,
    artifact_tag: str | None = None,
    settings: Settings | None = None,
) -> dict:
    """Gộp graph đã extract của docs và SQL sample; không gọi lại LLM."""
    app = settings or load_settings()
    artifacts = [
        load_graph(doc_id, artifact_tag=artifact_tag, settings=app)
        for doc_id in doc_ids
    ]
    nodes: dict[str, dict] = {}
    edges: dict[str, dict] = {}

    def extend_unique(target: list, values: list) -> None:
        for value in values:
            if value not in target:
                target.append(copy.deepcopy(value))

    for artifact in artifacts:
        for source in artifact["nodes"]:
            if source["id"] not in nodes:
                nodes[source["id"]] = copy.deepcopy(source)
                nodes[source["id"]].pop("community_id", None)
                continue
            node = nodes[source["id"]]
            descriptions = [item.strip() for item in (node.get("description", "") + " | " + source.get("description", "")).split("|") if item.strip()]
            node["description"] = " | ".join(dict.fromkeys(descriptions))
            for field in ("source_chunk_ids", "source_kinds", "evidence"):
                extend_unique(node.setdefault(field, []), source.get(field, []))
        for source in artifact["edges"]:
            if source["id"] not in edges:
                edges[source["id"]] = copy.deepcopy(source)
                continue
            edge = edges[source["id"]]
            descriptions = [item.strip() for item in (edge.get("description", "") + " | " + source.get("description", "")).split("|") if item.strip()]
            edge["description"] = " | ".join(dict.fromkeys(descriptions))
            for field in ("source_chunk_ids", "source_kinds", "evidence"):
                extend_unique(edge.setdefault(field, []), source.get(field, []))
            edge["weight"] = edge.get("weight", 1) + source.get("weight", 1)

    return {
        "doc_id": "+".join(doc_ids),
        "kind": "combined",
        "nodes": list(nodes.values()),
        "edges": list(edges.values()),
        "communities": [],
        "reports": [],
        "stats": {
            "source_artifacts": len(artifacts),
            "nodes": len(nodes),
            "edges": len(edges),
        },
    }


def _key(value: str) -> str:
    value = value.replace("\\_", "_").strip().upper()
    table_function = re.fullmatch(r"TABLE\s*\(\s*([^\)]+)\s*\)", value)
    if table_function:
        value = table_function.group(1)
    return re.sub(r"[^A-Z0-9_$#.]", "", value)


def load_cases(split: str, *, settings: Settings | None = None) -> list[dict]:
    import openpyxl

    app = settings or load_settings()
    cfg = app.eval
    source = app.path(app.paths.raw) / cfg.source
    if not source.exists():
        raise FileNotFoundError(source)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    if split not in workbook.sheetnames:
        raise ValueError(f"{source.name}: không có sheet '{split}'")
    rows = workbook[split].iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    required = (cfg.id_column, cfg.query_column, cfg.relevant_column)
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"sheet '{split}' thiếu cột: {', '.join(missing)}")
    positions = {name: headers.index(name) for name in required}
    cases = []
    for row in rows:
        case_id = str(row[positions[cfg.id_column]] or "").strip()
        query = str(row[positions[cfg.query_column]] or "").strip()
        relevant = str(row[positions[cfg.relevant_column]] or "").strip()
        if not case_id or not query:
            continue
        cases.append({
            "id": case_id,
            "query": query,
            "relevant": [item.strip() for item in relevant.split(cfg.separator) if item.strip()],
        })
    if not cases:
        raise ValueError(f"sheet '{split}' không có testcase hợp lệ")
    return cases


def evaluate(
    doc_id: str,
    *,
    knowledge_id: str,
    split: str | None = None,
    artifact: dict | None = None,
    output_tag: str | None = None,
    settings: Settings | None = None,
) -> dict:
    import networkx as nx

    app = settings or load_settings()
    selected_split = split or app.eval.split
    cases = load_cases(selected_split, settings=app)
    artifact = artifact or load_graph(doc_id, settings=app)
    nodes = artifact["nodes"]
    edges = artifact["edges"]
    table_nodes = {_key(node["name"]): node for node in nodes if node["type"] in {"table", "sql_function"}}
    graph = nx.Graph()
    graph.add_nodes_from(node["id"] for node in nodes)
    graph.add_edges_from((edge["source"], edge["target"]) for edge in edges)
    community_of = {node["id"]: node.get("community_id") for node in nodes}

    per_case = []
    all_gold: set[str] = set()
    all_matched: set[str] = set()
    path_lengths: list[int] = []
    for case in cases:
        relevant = {_key(item) for item in case["relevant"] if _key(item)}
        all_gold.update(relevant)
        matched = relevant & set(table_nodes)
        all_matched.update(matched)
        node_ids = [table_nodes[name]["id"] for name in matched]
        connected = len(node_ids) == len(relevant)
        distances = []
        if connected and len(node_ids) > 1:
            for index, source in enumerate(node_ids):
                for target in node_ids[index + 1:]:
                    if nx.has_path(graph, source, target):
                        distances.append(nx.shortest_path_length(graph, source, target))
                    else:
                        connected = False
        path_lengths.extend(distances)
        communities = {community_of.get(node_id) for node_id in node_ids if community_of.get(node_id)}
        per_case.append({
            "id": case["id"],
            "query": case["query"],
            "relevant": sorted(relevant),
            "matched": sorted(matched),
            "missing": sorted(relevant - matched),
            "table_recall": len(matched) / len(relevant) if relevant else 1.0,
            "complete": matched == relevant,
            "connected": connected,
            "same_community": bool(node_ids) and len(node_ids) == len(relevant) and len(communities) == 1,
        })

    count = len(per_case)
    metrics = {
        "cases": count,
        "unique_relevant_tables": len(all_gold),
        "matched_relevant_tables": len(all_matched),
        "table_coverage": len(all_matched) / len(all_gold) if all_gold else 1.0,
        "macro_table_recall": sum(case["table_recall"] for case in per_case) / count,
        "complete_case_rate": sum(case["complete"] for case in per_case) / count,
        "connected_case_rate": sum(case["connected"] for case in per_case) / count,
        "same_community_rate": sum(case["same_community"] for case in per_case) / count,
        "mean_shortest_path": sum(path_lengths) / len(path_lengths) if path_lengths else None,
    }
    parameters = {
        "knowledge_id": knowledge_id,
        "doc_id": doc_id,
        "split": selected_split,
        "chunk": app.chunk.model_dump(mode="json"),
        "graph": app.graph.model_dump(mode="json"),
        "embedding": app.embedding.model_dump(mode="json"),
        "retrieval": app.retrieval.model_dump(mode="json"),
        "graph_run": artifact.get("parameters", {}),
    }
    result = {
        "parameters": parameters,
        "graph_stats": artifact.get("stats", {}),
        "metrics": metrics,
        "per_case": per_case,
    }
    output_dir = app.path(app.paths.eval)
    output_dir.mkdir(parents=True, exist_ok=True)
    tag = f".{output_tag.strip().lower()}" if output_tag else ""
    if output_tag and not re.fullmatch(r"[a-z0-9][a-z0-9_-]*", output_tag.strip().lower()):
        raise ValueError("output_tag chỉ được chứa a-z, 0-9, '_' hoặc '-'")
    output = output_dir / f"{doc_id}{tag}.{selected_split}.graph_eval.json"
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {**result, "path": str(output)}
