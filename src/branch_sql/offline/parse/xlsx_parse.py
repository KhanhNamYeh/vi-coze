"""xlsx testcase -> markdown. Một trong ba loader của chặng `parse`.

Mỗi hàng Excel thành một mục có tiêu đề là mã testcase, bên dưới là ba nhãn:

    ## SQL_001

    query: 5 điểm bán hàng có số lượng thuê bao phát triển mới nhiều nhất...

    evidence: Điểm bán hàng được biểu diễn bằng SALE_CODE trong bảng...

    sql:

    ```sql
    SELECT * FROM (...)
    ```

Nhờ vậy SQL sample đi qua đúng ba chặng như tài liệu tri thức - `##` là ranh giới
đơn vị nên "một hàng = một chunk" là hệ quả của cấu trúc, không phải một luật cắt
riêng phải viết thêm.

SQL nằm trong khối fence chứ không thả trần: câu SQL thụt đầu dòng bằng khoảng
trắng, để trần thì CommonMark nuốt thành indented code block và ranh giới đoạn
lệch đi. Fence cũng giữ nguyên xuống dòng, thứ mà người đọc lại cần.

Sheet nào ứng với vai trò nào khai ở `parse.sheets` trong profile. Chỉ sheet `dev`
được dựng thành markdown để đem đi index; sheet `test` là bộ giữ kín, chỉ sinh
JSON qua `eval/gold_parse.py`.
"""

from __future__ import annotations

from pathlib import Path

from ...config import CFG

# Thứ tự cột trong sheet, khớp với `eval/gold_parse.py`.
COL_ID, COL_QUERY, COL_EVIDENCE, COL_SQL = 0, 1, 2, 3
TITLE = "Mẫu câu hỏi và câu SQL tương ứng"


def read_sheet(src: Path, sheet: str) -> list[tuple]:
    """Đọc một sheet, bỏ dòng đệm rỗng mà Excel hay để lại ở cuối."""
    import openpyxl

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(
            f"{src.name}: không có sheet '{sheet}' - có {', '.join(wb.sheetnames)}"
        )
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    # Dòng đầu là tiêu đề cột; dòng nào không có mã testcase thì bỏ.
    return [r for r in rows[1:] if r and str(r[COL_ID] or "").strip()]


def render(rows: list[tuple]) -> str:
    """Hàng Excel -> markdown có `#` nhóm / `##` mỗi mẫu."""
    out: list[str] = ["", f"# {TITLE}", ""]
    for row in rows:
        def cell(i: int) -> str:
            return str(row[i] or "").strip() if i < len(row) else ""

        out += ["", f"## {cell(COL_ID)}", ""]
        if query := cell(COL_QUERY):
            out += [f"query: {query}", ""]
        if evidence := cell(COL_EVIDENCE):
            out += [f"evidence: {evidence}", ""]
        if sql := cell(COL_SQL):
            out += ["sql:", "", "```sql", sql, "```", ""]
    return "\n".join(out)


def to_markdown(src: Path, *, role: str = "dev") -> str:
    """.xlsx -> markdown đã dựng cấu trúc, chưa làm sạch."""
    sheets = CFG.parse.sheets
    if not sheets:
        raise ValueError(
            f"{src.name}: profile chưa khai `parse.sheets` - không biết sheet nào "
            f"là bộ dev. Ví dụ: {{\"dev\": \"from schema\", \"test\": \"Sheet2\"}}"
        )
    if role not in sheets:
        raise ValueError(
            f"parse.sheets không có vai trò '{role}' - có {', '.join(sheets)}"
        )

    rows = read_sheet(src, sheets[role])
    if not rows:
        raise ValueError(f"{src.name}: sheet '{sheets[role]}' không có hàng nào")
    return render(rows)
