"""Text2SQL_testcase.xlsx -> ba artifact JSON. Bước nạp của chặng đo.

    uv run --extra eval python -m src.branch_sql.eval.gold_parse

Ở `eval/` chứ không ở `offline/parse/`: đây không phải parse tài liệu tri thức.
Nó đọc file testcase do người soạn, sinh ra tập đo, và `gold.py` ngay cạnh là
nơi tiêu thụ. Nó còn đối chiếu ngược với artifact của chặng `chunk`, tức là chạy
SAU pipeline offline chứ không nằm trong pipeline đó.

Nguồn nằm ở `RAW_DIR` vì đó là file người soạn, pipeline không ghi vào. Hai sheet
sinh ra ba file, vì sheet mẫu gánh hai vai trò khác nhau:

    parse.sheets["dev"]  -> PROCESSED_DIR/fewshot.chunks.json  tri thức, đem đi index
                         -> EVAL_DIR/dev.json                  bộ chỉnh tham số
    parse.sheets["test"] -> EVAL_DIR/test.json                 bộ báo cáo cuối

`fewshot` và `dev` cùng 18 dòng nhưng khác vai trò nên tách đường dẫn: một cái là
tri thức nạp vào hệ thống, một cái là câu hỏi dùng để chấm hệ thống. Cả hai sinh
từ cùng một sheet bằng cùng một lệnh nên không thể lệch nhau.

`test` chỉ đụng tới khi báo cáo. Chỉnh tham số dựa trên nó thì nó hết là tập giữ
kín, và con số cuối cùng thành con số đã nhìn trộm.

Trọn một hàng là một chunk, giữ đúng các cột của Excel, không thêm trường suy ra.
"""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from pathlib import Path

from ..config import CFG, EVAL_DIR, PROCESSED_DIR, RAW_DIR, rel
from .normalize import table_key

SOURCE_NAME = "Text2SQL_testcase.xlsx"

# Thứ tự cột trong sheet: Test Case ID | Query | Evidence | SQL | Relevant Chunks
COLS = 5


@dataclass(frozen=True)
class Output:
    sheet: str
    out_dir: Path
    out_name: str
    role: str


# Tên sheet lấy từ `parse.sheets` của profile, KHÔNG khai lại ở đây. Khai hai nơi
# thì đổi tên sheet trong Excel chỉ sửa được một chỗ, và chỗ còn lại hỏng im lặng
# - đã xảy ra thật khi sheet đổi từ "from schema"/"Sheet2" sang "dev"/"test".
_SHEETS = CFG.parse.sheets

OUTPUTS = (
    Output(_SHEETS["dev"], PROCESSED_DIR, "fewshot.chunks.json", "tri thức, đem đi index"),
    Output(_SHEETS["dev"], EVAL_DIR, "dev.json", "chỉnh tham số"),
    Output(_SHEETS["test"], EVAL_DIR, "test.json", "báo cáo cuối, chạm càng ít càng tốt"),
)


def read_sheet(src: Path, sheet: str) -> tuple[list[tuple], list[str]]:
    """Đọc sheet -> (các dòng hợp lệ, mã của dòng bị bỏ).

    Dòng phải có CẢ mã lẫn câu hỏi. Lọc theo mỗi cột mã là chưa đủ: người soạn
    hay đánh sẵn mã cho vài dòng rồi để trống nội dung, và những dòng đó lọt vào
    bộ đo thành câu hỏi rỗng - chấm ra 0 điểm tuyệt đối, kéo trung bình xuống mà
    trông như truy hồi kém chứ không như dữ liệu thiếu.
    """
    import openpyxl

    wb = openpyxl.load_workbook(src, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"{src.name}: không có sheet '{sheet}' (có: {', '.join(wb.sheetnames)})")

    rows, skipped = [], []
    for r in wb[sheet].iter_rows(min_row=2, values_only=True):
        if not r or not str(r[0] or "").strip():
            continue
        if not str(r[1] or "").strip():
            skipped.append(str(r[0]).strip())
            continue
        rows.append(r[:COLS])
    return rows, skipped


def build(src: Path, sheet: str) -> tuple[list[dict], list[str]]:
    """Một sheet -> (list chunk đúng 5 cột của Excel, mã dòng đã bỏ)."""
    rows, skipped = read_sheet(src, sheet)
    return [
        {
            "test_case_id": str(tid).strip(),
            "query": str(query).strip(),
            "evidence": str(evidence).strip(),
            "sql": str(sql).strip(),
            # ô Excel là chuỗi ngăn bằng "|", tách ra cho đúng dạng JSON
            "relevant_chunks": [t.strip() for t in str(relevant).split("|") if t.strip()],
        }
        for tid, query, evidence, sql, relevant in rows
    ], skipped


def write(records: list[dict], *, out_dir: Path, out_name: str) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    dst = out_dir / out_name
    dst.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return dst


def known_tables(chunks_path: Path) -> set[str]:
    """table_key của chunk schema, chỉ dùng để cảnh báo - không ghi vào file."""
    if not chunks_path.exists():
        return set()
    return {
        table_key(json.loads(line)["metadata"].get("table_name"))
        for line in chunks_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    }


def check(sheet: str, records: list[dict], known: set[str], seen: dict[str, str]) -> None:
    """Cảnh báo id trùng và tên bảng lạ. Chạy một lần mỗi sheet."""
    for r in records:
        tid = r["test_case_id"]
        if tid in seen and seen[tid] != sheet:
            print(f"     ! id '{tid}' có ở cả sheet '{seen[tid]}' và '{sheet}' - rò rỉ dev/test")
        seen[tid] = sheet

    if known:
        unknown = {t for r in records for t in r["relevant_chunks"] if table_key(t) not in known}
        for t in sorted(unknown):
            print(f"     ! '{t}' không khớp bảng nào trong chunk schema")


def main(argv: list[str]) -> int:
    src = RAW_DIR / SOURCE_NAME
    if not src.exists():
        print(f"không thấy {rel(src)}", file=sys.stderr)
        print(f"đặt {SOURCE_NAME} vào {rel(RAW_DIR)} trước", file=sys.stderr)
        return 1

    # Sai chính tả tên bảng chỉ lộ ra ở đây; lúc chấm điểm chỉ thấy recall thấp
    # mà không rõ nguyên nhân.
    known = known_tables(PROCESSED_DIR / "mo_ta_bang_bds_new.chunks.jsonl")
    cache: dict[str, list[dict]] = {}
    dropped: dict[str, list[str]] = {}
    seen: dict[str, str] = {}
    print(SOURCE_NAME)

    for out in OUTPUTS:
        if out.sheet not in cache:
            try:
                cache[out.sheet], dropped[out.sheet] = build(src, out.sheet)
            except ValueError as e:
                print(e, file=sys.stderr)
                return 1
            check(out.sheet, cache[out.sheet], known, seen)
            if ids := dropped[out.sheet]:
                print(f"  ! sheet '{out.sheet}': bỏ {len(ids)} dòng chỉ có mã, "
                      f"không có câu hỏi - {', '.join(ids)}")

        records = cache[out.sheet]
        dst = write(records, out_dir=out.out_dir, out_name=out.out_name)
        print(f"  [{out.sheet}] -> {rel(dst)}")
        print(f"     {len(records)} dòng | {out.role}")

    if not known:
        print("  ! chưa có chunk schema - bỏ qua bước đối chiếu tên bảng")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
